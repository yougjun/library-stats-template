"""
Excel Generator — Produces filled-in .xlsx workbooks using dynamic cell mappings.

Reads an Excel template workbook, populates it with data from the template-driven
data store using dynamic cell mappings, then writes the result to a temporary file.
"""

import openpyxl
from sqlalchemy.orm import Session
import tempfile
import uuid
import os
import logging

from app.config import Config
from app.services.excel.base import ExcelBaseGenerator
from app.mappings.dynamic_resolver import CellMappingResolver

logger = logging.getLogger(__name__)


class ExcelGenerator(ExcelBaseGenerator):
    def __init__(self, template_path: str, db: Session):
        super().__init__(db)
        self.template_path = template_path

    def generate(
        self,
        year_month: str,
        sheet_type: str = "all",
        **kwargs,
    ) -> str:
        logger.info(f"[ExcelGenerator.generate] entry: year_month={year_month}, sheet_type={sheet_type}")

        wb = openpyxl.load_workbook(self.template_path)

        self._fill_from_dynamic_mappings(wb, year_month, sheet_type)

        output_dir = tempfile.mkdtemp()
        output_path = os.path.join(output_dir, f"{uuid.uuid4().hex}.xlsx")
        wb.save(output_path)
        wb.close()

        logger.info(f"[ExcelGenerator.generate] exit: output={output_path}")
        return output_path

    def _fill_from_dynamic_mappings(self, wb, year_month: str, sheet_type: str):
        logger.debug(f"[_fill_from_dynamic_mappings] entry: year_month={year_month}")

        resolver = CellMappingResolver(self.db)
        mappings = resolver.get_mappings_for_template("current")

        if not mappings:
            logger.info("[_fill_from_dynamic_mappings] no dynamic mappings configured")
            return

        from app.models.template import CellData
        cell_data = (
            self.db.query(CellData)
            .filter(CellData.year_month == year_month)
            .all()
        )

        data_by_field = {}
        for cd in cell_data:
            data_by_field[cd.field_id] = cd.value

        filled = 0
        for cell_ref, field_id in mappings.items():
            if field_id not in data_by_field:
                continue

            parts = cell_ref.split("!")
            if len(parts) == 2:
                sheet_name, cell_addr = parts
            else:
                cell_addr = parts[0]
                sheet_name = wb.sheetnames[0] if wb.sheetnames else None

            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                try:
                    ws[cell_addr] = data_by_field[field_id]
                    filled += 1
                except Exception as e:
                    logger.warning(f"[_fill_from_dynamic_mappings] failed to write {cell_ref}: {e}")

        logger.info(f"[_fill_from_dynamic_mappings] exit: filled {filled} cells from {len(data_by_field)} data points")
