"""
Template Data Service — Merges saved CellData into Univer JSON and exports to xlsx.
Handles loading editor data with values, saving cell values, and Excel export.
"""

import logging
import re
from copy import deepcopy
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from sqlalchemy.orm import Session

from app.models.template import TemplateConfig, CellData

logger = logging.getLogger(__name__)


def cell_ref_to_indices(cell_ref: str) -> tuple[int, int]:
    match = re.match(r"^([A-Z]+)(\d+)$", cell_ref.upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    col_str, row_str = match.groups()
    row = int(row_str) - 1
    col = column_index_from_string(col_str) - 1
    return row, col


def indices_to_cell_ref(row: int, col: int) -> str:
    return f"{get_column_letter(col + 1)}{row + 1}"


def _find_sheet_id(univer_data: dict, sheet_name: str) -> Optional[str]:
    sheets = univer_data.get("sheets", {})
    for sheet_id, sheet_data in sheets.items():
        if sheet_data.get("name") == sheet_name:
            return sheet_id
    return None


def _cast_value(value: str, value_type: str) -> Any:
    if value is None:
        return None
    if value_type == "number":
        try:
            if "." in value:
                return float(value)
            return int(value)
        except (ValueError, TypeError):
            return value
    return value


def get_editor_data_with_values(db: Session, template_id: int, year_month: str) -> dict:
    logger.debug(f"[get_editor_data_with_values] entry: template={template_id}, month={year_month}")

    config = db.query(TemplateConfig).filter_by(id=template_id).first()
    if not config:
        raise ValueError(f"Template config not found: {template_id}")

    univer_data = deepcopy(config.univer_data)
    cell_roles = config.cell_roles or {}

    cells = db.query(CellData).filter_by(
        template_id=template_id,
        year_month=year_month,
    ).all()

    for cell in cells:
        sheet_id = _find_sheet_id(univer_data, cell.sheet_name)
        if not sheet_id:
            logger.warning(f"[get_editor_data_with_values] sheet not found: {cell.sheet_name}")
            continue

        try:
            row, col = cell_ref_to_indices(cell.cell_ref)
        except ValueError:
            logger.warning(f"[get_editor_data_with_values] invalid ref: {cell.cell_ref}")
            continue

        sheet_data = univer_data["sheets"][sheet_id]
        cell_data = sheet_data.setdefault("cellData", {})
        row_key = str(row) if isinstance(next(iter(cell_data), None), str) else row
        row_data = cell_data.setdefault(row_key, {})
        col_key = str(col) if isinstance(next(iter(row_data), None), str) else col

        existing = row_data.get(col_key, {})
        existing["v"] = _cast_value(cell.value, cell.value_type)
        if cell.value_type == "number":
            existing["t"] = 2
        else:
            existing["t"] = 1
        row_data[col_key] = existing

    logger.info(f"[get_editor_data_with_values] exit: merged {len(cells)} values")
    return {
        "univer_data": univer_data,
        "cell_roles": cell_roles,
    }


def save_cell_values(
    db: Session,
    template_id: int,
    year_month: str,
    cells: list[dict],
    updated_by: str = "",
) -> int:
    logger.debug(f"[save_cell_values] entry: template={template_id}, month={year_month}, count={len(cells)}")

    saved = 0
    for cell_info in cells:
        sheet_name = cell_info.get("sheet", "")
        cell_ref = cell_info.get("cell", "").upper()
        value = cell_info.get("value")
        value_type = cell_info.get("value_type", "number")

        if not sheet_name or not cell_ref:
            continue

        existing = db.query(CellData).filter_by(
            template_id=template_id,
            year_month=year_month,
            sheet_name=sheet_name,
            cell_ref=cell_ref,
        ).first()

        if existing:
            existing.value = str(value) if value is not None else None
            existing.value_type = value_type
            existing.updated_by = updated_by
        else:
            db.add(CellData(
                template_id=template_id,
                sheet_name=sheet_name,
                cell_ref=cell_ref,
                year_month=year_month,
                value=str(value) if value is not None else None,
                value_type=value_type,
                updated_by=updated_by,
            ))
        saved += 1

    db.commit()
    logger.info(f"[save_cell_values] exit: saved {saved} cells")
    return saved


def export_to_xlsx(
    db: Session,
    template_id: int,
    year_month: str,
    template_file_path: str,
    output_path: str,
) -> str:
    logger.debug(f"[export_to_xlsx] entry: template={template_id}, month={year_month}")

    config = db.query(TemplateConfig).filter_by(id=template_id).first()
    if not config:
        raise ValueError(f"Template config not found: {template_id}")

    wb = load_workbook(template_file_path)

    cells = db.query(CellData).filter_by(
        template_id=template_id,
        year_month=year_month,
    ).all()

    for cell in cells:
        if cell.sheet_name not in wb.sheetnames:
            logger.warning(f"[export_to_xlsx] sheet not found: {cell.sheet_name}")
            continue
        ws = wb[cell.sheet_name]
        ws[cell.cell_ref] = _cast_value(cell.value, cell.value_type)

    wb.save(output_path)
    wb.close()

    logger.info(f"[export_to_xlsx] exit: saved to {output_path}")
    return output_path
