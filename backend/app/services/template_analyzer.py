"""
Template Analyzer — Extracts structural metadata from Excel templates.
Identifies sheets, merged cells, headers (bold), data placeholders (value=0),
formula cells, and table regions bounded by borders.
"""

import logging
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def analyze_template(file_path: str) -> dict[str, Any]:
    logger.debug(f"[analyze_template] entry: path={file_path}")

    wb = load_workbook(file_path, data_only=False)
    sheets_info: list[dict[str, Any]] = []
    total_cells = 0
    formula_count = 0
    merged_count = 0

    for ws in wb.worksheets:
        sheet_result = _analyze_sheet(ws)
        sheets_info.append(sheet_result)
        total_cells += sheet_result["total_cells"]
        formula_count += len(sheet_result["formula_cells"])
        merged_count += len(sheet_result["merged_cells"])

    wb.close()

    result = {
        "sheets": sheets_info,
        "total_cells": total_cells,
        "formula_count": formula_count,
        "merged_count": merged_count,
    }

    logger.info(
        f"[analyze_template] exit: {len(sheets_info)} sheets, "
        f"{total_cells} cells, {formula_count} formulas, {merged_count} merges"
    )
    return result


def _analyze_sheet(ws) -> dict[str, Any]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    merged_cells = []
    for merged in ws.merged_cells.ranges:
        merged_cells.append({
            "range": str(merged),
            "start_row": merged.min_row,
            "start_col": merged.min_col,
            "end_row": merged.max_row,
            "end_col": merged.max_col,
        })

    headers: list[dict[str, Any]] = []
    data_cells: list[dict[str, Any]] = []
    formula_cells: list[dict[str, Any]] = []
    cell_count = 0

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None and not _cell_has_content(cell):
                continue

            cell_count += 1
            col_letter = get_column_letter(col_idx)
            ref = f"{col_letter}{row_idx}"

            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append({
                    "cell": ref,
                    "formula": cell.value,
                    "row": row_idx,
                    "col": col_idx,
                })
            elif cell.font and cell.font.bold and cell.value is not None:
                headers.append({
                    "cell": ref,
                    "value": str(cell.value),
                    "bold": True,
                    "row": row_idx,
                    "col": col_idx,
                })
            elif isinstance(cell.value, (int, float)) and cell.value == 0:
                data_cells.append({
                    "cell": ref,
                    "type": "number",
                    "value": 0,
                    "row": row_idx,
                    "col": col_idx,
                })

    tables = _detect_tables(ws, headers, data_cells, max_row, max_col)

    return {
        "name": ws.title,
        "dimensions": {"rows": max_row, "cols": max_col},
        "merged_cells": merged_cells,
        "headers": headers,
        "data_cells": data_cells,
        "formula_cells": formula_cells,
        "tables": tables,
        "total_cells": cell_count,
    }


def _cell_has_content(cell) -> bool:
    if cell.font and cell.font.bold:
        return True
    if cell.fill and cell.fill.fill_type == "solid":
        return True
    border = cell.border
    if border and any([
        border.top and border.top.style,
        border.bottom and border.bottom.style,
        border.left and border.left.style,
        border.right and border.right.style,
    ]):
        return True
    return False


def _detect_tables(ws, headers, data_cells, max_row, max_col) -> list[dict[str, Any]]:
    header_rows = {}
    for h in headers:
        row = h["row"]
        if row not in header_rows:
            header_rows[row] = []
        header_rows[row].append(h)

    data_by_row: dict[int, list] = {}
    for d in data_cells:
        row = d["row"]
        if row not in data_by_row:
            data_by_row[row] = []
        data_by_row[row].append(d)

    tables: list[dict[str, Any]] = []
    table_idx = 0

    sorted_header_rows = sorted(header_rows.keys())
    for header_row in sorted_header_rows:
        h_cells = header_rows[header_row]
        if len(h_cells) < 2:
            continue

        data_rows = []
        for r in range(header_row + 1, min(header_row + 30, max_row + 1)):
            if r in data_by_row:
                data_rows.append(r)
            elif r in header_rows:
                break

        if not data_rows:
            continue

        cols = sorted(set(h["col"] for h in h_cells))
        col_letters = [get_column_letter(c) for c in cols]

        header_map = {}
        for h in h_cells:
            header_map[get_column_letter(h["col"])] = str(h["value"])

        tables.append({
            "name": f"table_{table_idx}",
            "header_row": header_row,
            "data_rows": data_rows,
            "columns": col_letters,
            "headers": header_map,
        })
        table_idx += 1

    return tables


def auto_detect_cell_roles(structure: dict) -> dict[str, str]:
    roles: dict[str, str] = {}

    for sheet in structure.get("sheets", []):
        sheet_name = sheet["name"]

        for h in sheet.get("headers", []):
            key = f"{sheet_name}!{h['cell']}"
            roles[key] = "header"

        for d in sheet.get("data_cells", []):
            key = f"{sheet_name}!{d['cell']}"
            roles[key] = "input"

        for f in sheet.get("formula_cells", []):
            key = f"{sheet_name}!{f['cell']}"
            roles[key] = "computed"

    return roles
