"""
Template Converter — Bidirectional conversion between openpyxl and Univer IWorkbookData.

Converts .xlsx files to Univer JSON format for the in-browser editor,
and converts Univer JSON back to .xlsx for persistence.
"""

import logging
import uuid
from typing import Any, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

BORDER_STYLE_MAP = {
    "thin": 1,
    "medium": 2,
    "thick": 3,
    "dashed": 4,
    "dotted": 5,
    "double": 6,
    "hair": 7,
    "mediumDashed": 8,
    "dashDot": 9,
    "mediumDashDot": 10,
    "dashDotDot": 11,
    "mediumDashDotDot": 12,
    "slantDashDot": 13,
}

BORDER_STYLE_REVERSE = {v: k for k, v in BORDER_STYLE_MAP.items()}

H_ALIGN_MAP = {"left": 0, "center": 1, "right": 2, "general": 0}
V_ALIGN_MAP = {"top": 0, "center": 1, "bottom": 2}
H_ALIGN_REVERSE = {0: "left", 1: "center", 2: "right"}
V_ALIGN_REVERSE = {0: "top", 1: "center", 2: "bottom"}


def _color_to_rgb(color) -> Optional[str]:
    if color is None or color.type is None:
        return None
    if color.type == "rgb" and color.rgb and color.rgb != "00000000":
        rgb = str(color.rgb)
        if len(rgb) == 8:
            rgb = rgb[2:]
        return f"#{rgb}"
    if color.type == "indexed":
        return None
    if color.type == "theme":
        return None
    return None


def _border_side_to_univer(side: Optional[Side]) -> Optional[dict]:
    if side is None or side.style is None:
        return None
    result: dict[str, Any] = {"s": BORDER_STYLE_MAP.get(side.style, 1)}
    color = _color_to_rgb(side.color) if side.color else None
    if color:
        result["cl"] = {"rgb": color}
    return result


def _style_to_univer(cell: Cell) -> dict:
    s: dict[str, Any] = {}

    font = cell.font
    if font:
        if font.bold:
            s["bl"] = 1
        if font.italic:
            s["it"] = 1
        if font.underline and font.underline != "none":
            s["ul"] = {"s": 1}
        if font.strikethrough:
            s["st"] = {"s": 1}
        if font.name:
            s["ff"] = font.name
        if font.size:
            s["fs"] = font.size
        color = _color_to_rgb(font.color) if font.color else None
        if color:
            s["cl"] = {"rgb": color}

    fill = cell.fill
    if fill and fill.fill_type == "solid":
        fg_color = _color_to_rgb(fill.fgColor) if fill.fgColor else None
        if fg_color:
            s["bg"] = {"rgb": fg_color}

    border = cell.border
    if border:
        bd: dict[str, Any] = {}
        for key, attr in [("t", "top"), ("b", "bottom"), ("l", "left"), ("r", "right")]:
            side_data = _border_side_to_univer(getattr(border, attr, None))
            if side_data:
                bd[key] = side_data
        if bd:
            s["bd"] = bd

    alignment = cell.alignment
    if alignment:
        if alignment.horizontal and alignment.horizontal in H_ALIGN_MAP:
            s["ht"] = H_ALIGN_MAP[alignment.horizontal]
        if alignment.vertical and alignment.vertical in V_ALIGN_MAP:
            s["vt"] = V_ALIGN_MAP[alignment.vertical]
        if alignment.wrap_text:
            s["tb"] = 2
        if alignment.text_rotation:
            s["tr"] = {"a": alignment.text_rotation}

    if cell.number_format and cell.number_format != "General":
        s["n"] = {"pattern": cell.number_format}

    return s


def _cell_to_univer(cell: Cell, style_index: int, cached_value=None) -> dict:
    result: dict[str, Any] = {}

    if cell.value is not None:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            result["f"] = cell.value
            result["t"] = 2
            if cached_value is not None:
                result["v"] = cached_value
        elif isinstance(cell.value, (int, float)):
            result["v"] = cell.value
            result["t"] = 2
        elif isinstance(cell.value, bool):
            result["v"] = 1 if cell.value else 0
            result["t"] = 1
        else:
            result["v"] = str(cell.value)
            result["t"] = 1

    if style_index >= 0:
        result["s"] = style_index

    return result


def xlsx_to_univer(file_path: str) -> dict:
    logger.debug(f"[xlsx_to_univer] entry: path={file_path}")

    wb = load_workbook(file_path, data_only=False)
    wb_cached = load_workbook(file_path, data_only=True)
    sheets: dict[str, Any] = {}
    sheet_order: list[str] = []
    styles: dict[str, Any] = {}
    style_cache: dict[str, int] = {}
    style_counter = 0

    for idx, ws_name in enumerate(wb.sheetnames):
        ws = wb[ws_name]
        ws_cached = wb_cached[ws_name]
        sheet_id = f"sheet_{idx}"
        sheet_order.append(sheet_id)

        cell_data: dict[int, dict[int, Any]] = {}
        row_data: dict[int, Any] = {}
        column_data: dict[int, Any] = {}

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None and not _has_style(cell):
                    continue

                style = _style_to_univer(cell)
                style_key = str(sorted(style.items())) if style else ""
                s_idx = -1
                if style:
                    if style_key in style_cache:
                        s_idx = style_cache[style_key]
                    else:
                        s_idx = style_counter
                        style_cache[style_key] = s_idx
                        styles[str(s_idx)] = style
                        style_counter += 1

                cached_val = ws_cached.cell(row=row_idx, column=col_idx).value
                cell_dict = _cell_to_univer(cell, s_idx, cached_value=cached_val)
                if cell_dict:
                    r = row_idx - 1
                    c = col_idx - 1
                    if r not in cell_data:
                        cell_data[r] = {}
                    cell_data[r][c] = cell_dict

        for r in range(1, max_row + 1):
            dim = ws.row_dimensions.get(r)
            if dim and dim.height is not None:
                row_data[r - 1] = {"h": dim.height, "hd": 0}

        for c in range(1, max_col + 1):
            col_letter = get_column_letter(c)
            dim = ws.column_dimensions.get(col_letter)
            if dim and dim.width is not None:
                column_data[c - 1] = {"w": dim.width * 7.5, "hd": 0}

        merge_data = []
        for merged in ws.merged_cells.ranges:
            merge_data.append({
                "startRow": merged.min_row - 1,
                "endRow": merged.max_row - 1,
                "startColumn": merged.min_col - 1,
                "endColumn": merged.max_col - 1,
            })

        sheets[sheet_id] = {
            "id": sheet_id,
            "name": ws_name,
            "rowCount": max(max_row, 100),
            "columnCount": max(max_col, 26),
            "cellData": cell_data,
            "rowData": row_data,
            "columnData": column_data,
            "mergeData": merge_data,
            "tabColor": "",
            "defaultRowHeight": 20,
            "defaultColumnWidth": 73,
        }

    wb.close()
    wb_cached.close()

    result = {
        "id": str(uuid.uuid4()),
        "name": "Template",
        "appVersion": "0.0.0",
        "locale": "koKR",
        "styles": styles,
        "sheetOrder": sheet_order,
        "sheets": sheets,
    }

    logger.info(f"[xlsx_to_univer] exit: {len(sheet_order)} sheets, {len(styles)} styles")
    return result


def _has_style(cell: Cell) -> bool:
    if cell.font and (cell.font.bold or cell.font.italic or cell.font.name != "Calibri"):
        return True
    if cell.fill and cell.fill.fill_type == "solid":
        return True
    if cell.border and any([
        cell.border.top and cell.border.top.style,
        cell.border.bottom and cell.border.bottom.style,
        cell.border.left and cell.border.left.style,
        cell.border.right and cell.border.right.style,
    ]):
        return True
    if cell.alignment and (cell.alignment.horizontal or cell.alignment.vertical or cell.alignment.wrap_text):
        return True
    if cell.number_format and cell.number_format != "General":
        return True
    return False


def univer_to_xlsx(data: dict, output_path: str) -> str:
    logger.debug(f"[univer_to_xlsx] entry: output={output_path}")

    wb = Workbook()
    wb.remove(wb.active)

    styles = data.get("styles", {})
    sheet_order = data.get("sheetOrder", [])
    sheets = data.get("sheets", {})

    for sheet_id in sheet_order:
        sheet_data = sheets.get(sheet_id)
        if not sheet_data:
            continue

        ws = wb.create_sheet(title=sheet_data.get("name", sheet_id))
        cell_data = sheet_data.get("cellData", {})

        for row_key, row_cells in cell_data.items():
            r = int(row_key) + 1
            for col_key, cell_info in row_cells.items():
                c = int(col_key) + 1
                cell = ws.cell(row=r, column=c)

                f = cell_info.get("f")
                if f and isinstance(f, str) and f.startswith("="):
                    cell.value = f
                else:
                    v = cell_info.get("v")
                    t = cell_info.get("t")
                    if v is not None:
                        if t == 2:
                            try:
                                cell.value = float(v) if isinstance(v, str) else v
                            except (ValueError, TypeError):
                                cell.value = v
                        else:
                            cell.value = v

                s_idx = cell_info.get("s")
                if s_idx is not None:
                    style = styles.get(str(s_idx), {})
                    _apply_univer_style(cell, style)

        row_data = sheet_data.get("rowData", {})
        for row_key, row_info in row_data.items():
            r = int(row_key) + 1
            h = row_info.get("h")
            if h is not None:
                ws.row_dimensions[r].height = h

        column_data = sheet_data.get("columnData", {})
        for col_key, col_info in column_data.items():
            c = int(col_key) + 1
            w = col_info.get("w")
            if w is not None:
                col_letter = get_column_letter(c)
                ws.column_dimensions[col_letter].width = w / 7.5

        merge_data = sheet_data.get("mergeData", [])
        for m in merge_data:
            ws.merge_cells(
                start_row=m["startRow"] + 1,
                start_column=m["startColumn"] + 1,
                end_row=m["endRow"] + 1,
                end_column=m["endColumn"] + 1,
            )

    wb.save(output_path)
    wb.close()
    logger.info(f"[univer_to_xlsx] exit: saved to {output_path}")
    return output_path


def _apply_univer_style(cell: Cell, style: dict) -> None:
    font_kwargs: dict[str, Any] = {}
    if style.get("bl"):
        font_kwargs["bold"] = True
    if style.get("it"):
        font_kwargs["italic"] = True
    if style.get("ul"):
        font_kwargs["underline"] = "single"
    if style.get("st"):
        font_kwargs["strikethrough"] = True
    if style.get("ff"):
        font_kwargs["name"] = style["ff"]
    if style.get("fs"):
        font_kwargs["size"] = style["fs"]
    cl = style.get("cl")
    if cl and isinstance(cl, dict):
        rgb = cl.get("rgb", "").lstrip("#")
        if rgb:
            font_kwargs["color"] = rgb
    if font_kwargs:
        cell.font = Font(**font_kwargs)

    bg = style.get("bg")
    if bg and isinstance(bg, dict):
        rgb = bg.get("rgb", "").lstrip("#")
        if rgb:
            cell.fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")

    bd = style.get("bd")
    if bd and isinstance(bd, dict):
        sides = {}
        for key, attr in [("t", "top"), ("b", "bottom"), ("l", "left"), ("r", "right")]:
            side_data = bd.get(key)
            if side_data:
                sides[attr] = _univer_to_border_side(side_data)
            else:
                sides[attr] = Side()
        cell.border = Border(**sides)

    align_kwargs: dict[str, Any] = {}
    if "ht" in style:
        align_kwargs["horizontal"] = H_ALIGN_REVERSE.get(style["ht"], "left")
    if "vt" in style:
        align_kwargs["vertical"] = V_ALIGN_REVERSE.get(style["vt"], "top")
    if style.get("tb") == 2:
        align_kwargs["wrap_text"] = True
    tr = style.get("tr")
    if tr and isinstance(tr, dict) and "a" in tr:
        align_kwargs["text_rotation"] = tr["a"]
    if align_kwargs:
        cell.alignment = Alignment(**align_kwargs)

    n = style.get("n")
    if n and isinstance(n, dict) and "pattern" in n:
        cell.number_format = n["pattern"]


def _univer_to_border_side(data: dict) -> Side:
    style_num = data.get("s", 1)
    style_name = BORDER_STYLE_REVERSE.get(style_num, "thin")
    color = None
    cl = data.get("cl")
    if cl and isinstance(cl, dict):
        color = cl.get("rgb", "").lstrip("#")
    if color:
        return Side(style=style_name, color=color)
    return Side(style=style_name)
